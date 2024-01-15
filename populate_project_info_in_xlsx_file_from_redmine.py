import argparse
import requests
import yaml
import openpyxl
import pdb
from pprint import pprint




def get_custom_field(issue, field_name):
    """ Get custom field value from issue """

    # get field value
    field_value = [ field['value'] for field in issue['custom_fields'] if field['name']==field_name ]
    if len(field_value) > 0:
        field_value = field_value[0]
    else:
        field_value = ''

    return field_value

def fetch_redmine_users(redmine_url, api_key):
    # Make a request to the Redmine API to fetch all users
    users = {}
    offset = 0
    limit = 100
    total_count = float('inf')

    headers = {"X-Redmine-API-Key": api_key}

    while offset < total_count:
        params = {"offset": offset, "limit": limit}
        response = requests.get(f"{redmine_url}/users.json", headers=headers, params=params)

        if response.status_code == 200:
            data = response.json()
            total_count = data["total_count"]

            for user in data["users"]:
                user_id = user["id"]
                user_name = user["firstname"] + " " + user["lastname"]
                users[user_id] = user_name

            offset += limit
        else:
            break

    return users

def fetch_redmine_ticket(redmine_url, api_key, ticket_id):
    # Make a request to the Redmine API to fetch the ticket information
    headers = {"X-Redmine-API-Key": api_key}
    response = requests.get(f"{redmine_url}/issues/{ticket_id}.json", headers=headers)
    if response.status_code == 200:
        return response.json()["issue"]
    else:
        return None

def populate_xlsx_file(redmine_url, api_key, xlsx_file_path):

    # Fetch all users from the Redmine API
    redmine_users = fetch_redmine_users(redmine_url, api_key)

    # Open the existing xlsx file
    workbook = openpyxl.load_workbook(xlsx_file_path)
    worksheet = workbook["Projects Active"]

    # Insert two new columns to the left of column A
    worksheet.insert_cols(3, 3)

    # Write the column names in the new columns
    worksheet.cell(row=2, column=3, value="Assignee")
    worksheet.cell(row=2, column=4, value="Coordinator")
    worksheet.cell(row=2, column=5, value="Project Name")

    # Find the column index for "Project ID"
    project_id_column = None
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col).value == "Project ID":
            project_id_column = col
            break

    # save header formatting
    header_format = worksheet.cell(row=2, column=project_id_column).font.copy()

    # apply header formatting to new columns
    for col in range(3, 6):
        worksheet.cell(row=2, column=col).font = header_format

    # Iterate through each row in the column with "Project ID"
    for row in range(3, worksheet.max_row + 1):
        project_id = worksheet.cell(row=row, column=project_id_column).value

        # Check if project_id can be converted to an integer
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            continue

        print(f"Fetching Redmine ticket for project ID {project_id}...")

        # Fetch the Redmine ticket information
        ticket = fetch_redmine_ticket(redmine_url, api_key, project_id)

        if ticket:

            # Get the coordinator name
            coordinator_id = get_custom_field(ticket, 'Coordinator')
            try:
                coordinator = redmine_users[int(coordinator_id)]
            except (KeyError, ValueError):
                coordinator = ""

            # Write the ticket information to the new columns
            try:
                worksheet.cell(row=row, column=3, value=ticket["assigned_to"]["name"])
            except KeyError:
                worksheet.cell(row=row, column=3, value="")
            worksheet.cell(row=row, column=4, value=coordinator)
            worksheet.cell(row=row, column=5, value=ticket["subject"])

            # Shift the existing columns to the right
#            for col in range(worksheet.max_column, 3, -1):
#                worksheet.cell(row=row, column=col).value = worksheet.cell(row=row, column=col - 2).value

    # Save the modified xlsx file
    workbook.save(xlsx_file_path)

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description="Populate an xlsx file with data from the Redmine API")
    parser.add_argument("redmine_credentials", help="Path to the YAML file containing Redmine API key")
    parser.add_argument("xlsx_file_path", help="Path to the xlsx file")
    args = parser.parse_args()

    # Read the Redmine URL and API key from the YAML file
    with open(args.redmine_credentials, "r") as file:
        config = yaml.safe_load(file)

    redmine_url = config["url"]
    api_key = config["api_key"]

    # Populate the xlsx file with data from the Redmine API
    populate_xlsx_file(redmine_url, api_key, args.xlsx_file_path)

if __name__ == "__main__":
    main()
