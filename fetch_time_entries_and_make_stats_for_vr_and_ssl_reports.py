#!/uisr/bin/env python3
# -*- coding: utf-8 -*-
import pdb
from pprint import pprint

import argparse
import requests
import yaml
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font
import re

from collections import defaultdict
def nested_dict():
        return defaultdict(nested_dict)



def uni_shortname2longname(uni):

    # define translation table
    translation = {
        'Chalmers'                          : 'Chalmers University of Technology',
        'KI'                                : 'Karolinska Institutet',
        'KTH'                               : 'KTH Royal Institute of Technology',
        'LiU'                               : 'Linköping University',
        'LU'                                : 'Lund University',
        'SU'                                : 'Stockholm University',
        'SLU'                               : 'Swedish University of Agricultural Sciences',
        'UmU'                               : 'Umeå University',
        'GU'                                : 'University of Gothenburg',
        'UU'                                : 'Uppsala University',
        'NRM'                               : 'Naturhistoriska Riksmuséet',
        'LNU'                               : 'Linnaeus University',
        'Örebro University'                 : 'Örebro University',
        'Other Swedish University'          : 'Other Swedish University',
        'Other Swedish organization'        : 'Other Swedish organization',
        'Healthcare'                        : 'Healthcare',
        'Industry'                          : 'Industry',
        'International University'          : 'International University',
        'Other international organization'  : 'Other international organization',
    }

    if uni not in translation:
        # should we look at PIs email to determin this_
        print(f"WARNING: uni not in translation list, {uni}")

    # return translation if it exists, otherwise return it untranslated
    return translation.get(uni, uni)




def uni_from_pi_email(email):
    """
    Guess the project's university based on the PIs email domain.
    """

    domain = email.split(".")[-2]
    pass




def fetch_time_entries(start_date, end_date, url, api_key):
    """
    Fetches the time entries within the specified date range and project ID.

    Args:
        start_date (str): Start date in format 'YYYY-MM-DD'.
        end_date (str): End date in format 'YYYY-MM-DD'.
        project_id (int): Project ID.
        url (str): Redmine URL.
        api_key (str): Redmine API key.

    Returns:
        set: Set of unique issue IDs.
    """
    params = {
        'key': api_key,
        'spent_on': f'><{start_date}|{end_date}',
#        'project_id': project_id,
        'limit': 100,
        'offset': 0
    }
    issue_ids = nested_dict()

    response = requests.get(f'{url}/time_entries.json', params=params)
    response.raise_for_status()
    data = response.json()

    total_count = data['total_count']

    # Fetch time entries in batches
    while params['offset'] < total_count:
        response = requests.get(f'{url}/time_entries.json', params=params)
        response.raise_for_status()
        data = response.json()

        time_entries = data['time_entries']

        for entry in time_entries:
            try:
                issue_ids[entry['issue']['id']][entry['activity']['name']] += entry['hours']

            except:
                issue_ids[entry['issue']['id']][entry['activity']['name']] = entry['hours']

        params['offset'] += params['limit']

        # Calculate progress percentage
        progress = min(params['offset'], total_count) / total_count * 100
        print(f'Fetching time entries: {progress:.2f}% complete', end='\r')

    print('Fetching time entries: 100% complete')
    
    return issue_ids

def fetch_issue_details(issue_ids, url, api_key):
    """
    Fetches the detailed information about each issue.

    Args:
        issue_ids (set): Set of unique issue IDs.
        url (str): Redmine URL.
        api_key (str): Redmine API key.

    Returns:
        list: List of issue details.
    """
    issue_details = []

    print('Fetching issue details:')
    for i, issue_id in enumerate(issue_ids, 1):
        response = requests.get(f'{url}/issues/{issue_id}.json', params={'key': api_key})
        response.raise_for_status()
        data = response.json()

        # Calculate progress percentage
        progress = i / len(issue_ids) * 100
        print(f'Progress: {progress:.2f}%', end='\r')

        # skip issues from other projects than defined below#        if data['issue'] ['project']['name'] not in ['National Bioinformatics Support', ] and not re.search('^Round ', data['issue'] ['project']['name']):
        # should we keep all and filter later instead?
        if data['issue'] ['project']['name'] not in ['National Bioinformatics Support', ] and not re.search('^Round ', data['issue'] ['project']['name']):
            continue
       
        data['issue']['spent_per_activity'] = dict(issue_ids[issue_id])
        issue_details.append(data['issue'])


    print('Progress: 100%')

    return issue_details

def generate_statistics(issue_details):
    """
    Generates statistics for each issue.

    Args:
        issue_details (list): List of issue details.

    Returns:
        list: List of dictionaries containing statistics for each issue.
    """
    statistics = []

    for issue in issue_details:

        #pdb.set_trace()

        # get custom fields data
        cfs = {}
        for cf in issue['custom_fields']:

            if cf['name'] == 'Principal Investigator':
                try:
                    cfs['pi_fullname']  = cf['value']
                    cfs['pi_firstname'] = " ".join(cfs['pi_fullname'].split()[:-1])
                    cfs['pi_lastname']  = cfs['pi_fullname'].split()[-1]
                except:
                    cfs['pi_firstname'] = ''
                    cfs['pi_lastname']  = ''


            elif cf['name'] == 'PI e-mail':
                cfs['pi_email']     = cf['value']

            elif cf['name'] == 'Organization':
                cfs['org']          = uni_shortname2longname(cf['value'])

            elif cf['name'] == 'Subject':
                cfs['subject']      = cf['value']

            elif cf['name'] == 'SCB Subject Code':
                cfs['scb']          = cf['value']

            elif cf['name'] == 'Funding':
                cfs['funding']      = cf['value']

            elif cf['name'] == 'Publication(s)':
                cfs['publications'] = cf['value']

            elif cf['name'] == 'PI Gender':
                cfs['pi_gender']    = cf['value']

            elif cf['name'] == 'WABI ID':
                cfs['wabi_id']      = cf['value']

        # pdb.set_trace()
        statistics.append({
            'Project ID'        : issue['id'],
            'PI first name'     : cfs.get('pi_firstname',''),
            'PI last name'      : cfs.get('pi_lastname',''),
            'email'             : cfs.get('pi_email', ''),
            'Organization'      : cfs.get('org',''),
            'SCB Subject Code'  : cfs.get('scb',''),
            'Sex'               : cfs.get('pi_gender',''),
            'Subject'           : cfs.get('subject',''),
            'LTS project ID'    : cfs.get('wabi_id',''),
            'Publications'      : cfs.get('publications',''),
            'Funding'           : cfs.get('funding','')
        })

    return statistics




def get_custom_field(issue, field_name):


    # get field value
    field_value = [ field['value'] for field in issue['custom_fields'] if field['name']==field_name ]
    if len(field_value) > 0:
        field_value = field_value[0]
    else:
        field_value = ''

    return field_value




def save_issues_as_excel(issue_details, output_path):
    """
    Saves the issues as an Excel file and makes statistics as well.

    Args:
        issue_details (list): List of dictionaries containing issues.
        output_path (str): Path to save the Excel file.
    """
    workbook = Workbook()
    pl_sheet = workbook.active
    pl_sheet.title = "Project list"

    bold_font = Font(bold=True)

    # Write headers
#    headers = list(statistics[0].keys())
    headers = ['Project ID', 'PI first name', 'PI last name', 'email', 'Organization', 'SCB Subject Code', 'Sex', 'Tracker', 'LTS project ID', 'Publications', 'Funding', 'Spent time this period', 'Spent time total']
    for col_num, header in enumerate(headers, 1):
        pl_sheet.cell(row=1, column=col_num, value=header)
        pl_sheet.cell(row=1, column=col_num).font = bold_font

    # Write data rows
    for row_num, issue in enumerate(issue_details, 2):

            # get PI first and last name
            pi_name = get_custom_field(issue, 'Principal Investigator')
            pi_name_split = pi_name.split(' ')
            pi_last_name  = pi_name_split[-1]
            pi_first_name = " ".join(pi_name_split[:-1])

            # summarize the hours spent the requested period
            time_spent_this_period = sum([ hours for hours in issue['spent_per_activity'].values() ])

            #pdb.set_trace()

            # print values
            pl_sheet.cell(row=row_num, column=1, value=issue.get('id',''))
            pl_sheet.cell(row=row_num, column=2, value=pi_first_name)
            pl_sheet.cell(row=row_num, column=3, value=pi_last_name)
            pl_sheet.cell(row=row_num, column=4, value=get_custom_field(issue, 'PI e-mail'))
            pl_sheet.cell(row=row_num, column=5, value=get_custom_field(issue, 'Organization'))
            pl_sheet.cell(row=row_num, column=6, value=get_custom_field(issue, 'SCB Subject Code'))
            pl_sheet.cell(row=row_num, column=7, value=get_custom_field(issue, 'PI Gender'))
            pl_sheet.cell(row=row_num, column=8, value=issue.get('tracker',{}).get('name',''))
            pl_sheet.cell(row=row_num, column=9, value=get_custom_field(issue, 'WABI ID'))
            pl_sheet.cell(row=row_num, column=10, value=get_custom_field(issue, 'Publication(s)'))
            pl_sheet.cell(row=row_num, column=11, value=get_custom_field(issue, 'Funding'))
            pl_sheet.cell(row=row_num, column=12, value=time_spent_this_period)
            pl_sheet.cell(row=row_num, column=13, value=issue.get('spent_hours',''))


    ### do other statistics

    # PIs per uni
    if 1:
        ppo_sheet = workbook.create_sheet("Projects per org")

        # print headers
        ppo_sheet.cell(row=1, column=1, value="Organization")
        ppo_sheet.cell(row=1, column=1).font = bold_font
        ppo_sheet.cell(row=1, column=2, value="#")
        ppo_sheet.cell(row=1, column=2).font = bold_font

        # print the UNIQUE function to get all org names
        #ppo_sheet.cell(row=2, column=1, value="=_xlfn.UNIQUE('Project list'!E2:'Project list'!E1000)")
        ppo_sheet.cell(row=2, column=1, value=ArrayFormula(f"A2:A50", f"=_xlfn.UNIQUE('Project list'!E2:'Project list'!E1000)"))

        # print the counting function
        for i in range(2,20):
            #pdb.set_trace()
            ppo_sheet.cell(row=i, column=2, value=f"=_xlfn.COUNTIF('Project list'!$E$2:'Project list'!$E$1000, A{i})")
            #ppo_sheet.cell(row=i, column=2, value=f"=SUM(4,7)")
















    workbook.save(output_path)
    print(f'Statistics saved as {output_path}')

def main():
    parser = argparse.ArgumentParser(description='Fetch Redmine time entries between two dates')
    parser.add_argument('-s', '--start-date', required=True, help='Start date in YYYY-MM-DD format')
    parser.add_argument('-e', '--end-date', required=True, help='End date in YYYY-MM-DD format')
#    parser.add_argument('-p', '--project-id', required=True, type=str, help='Project ID')
    parser.add_argument('-c', '--config', required=True, help='Config file path')
    parser.add_argument('-o', '--output', required=True, help='Output file path')
    args = parser.parse_args()

    with open(args.config) as f:
        config = yaml.safe_load(f)

    issue_ids       = fetch_time_entries(args.start_date, args.end_date, config['url'], config['api_key'])
    issue_details   = fetch_issue_details(issue_ids, config['url'], config['api_key'])
    #statistics      = generate_statistics(issue_details)
    save_issues_as_excel(issue_details, args.output)

if __name__ == '__main__':
    main()
